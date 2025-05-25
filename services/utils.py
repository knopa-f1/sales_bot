import datetime
import os

from openpyxl.reader.excel import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook

from config_data.config import config_settings
from db.requests.carts import get_cart_items


def format_cart_message(cart_items, total) -> str:
    if not cart_items:
        return "🛒 Ваша корзина пуста."

    lines = ["🛒 <b>Ваша корзина:</b>\n"]

    for item in cart_items:
        name = item.product.name
        price = float(item.product.price)
        count = item.count
        subtotal = price * count
        lines.append(f"• {name} — {count} x {price:.2f} ₽ = <b>{subtotal:.2f} ₽</b>")

    lines.append(f"\n<b>Итого:</b> {total:.2f} ₽")

    return "\n".join(lines)

def format_order_confirmation_message(order) -> str:
    return (
        f"✅ <b>Заказ №{order.id}</b> успешно оформлен!\n\n"
        f"💰 Сумма: <b>{order.amount:.2f} ₽</b>\n"
        f"📦 Доставка по адресу:\n{order.address}\n"
    )

async def append_payment_to_excel(order_id: int, user_id: int, username: str, amount: float,
                                  currency: str, provider_tx_id: str, timestamp: datetime):
    file_path = config_settings.report_path

    headers = ["Order ID", "User ID", "Username", "Amount", "Currency", "Transaction ID", "Date"]

    if not os.path.exists(file_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Payments"
        ws.append(headers)
    else:
        wb = load_workbook(file_path)
        ws = wb.active

    ws.append([
        order_id,
        user_id,
        username,
        f"{amount:.2f}",
        currency,
        provider_tx_id,
        timestamp.strftime("%Y-%m-%d %H:%M:%S")
    ])

    for col in ws.columns:
        max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max_len + 2

    wb.save(file_path)
